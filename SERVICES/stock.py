import os
import pandas as pd
from openpyxl import load_workbook


CARPETA_STOCK = "ARCHIVOS/STOCK"


# ============================================================
# BUSCAR ÚLTIMO STOCK
# ============================================================

def buscar_ultimo_stock():

    archivos = []

    if not os.path.exists(CARPETA_STOCK):
        raise FileNotFoundError(
            f"No existe la carpeta {CARPETA_STOCK}."
        )

    for archivo in os.listdir(CARPETA_STOCK):

        ruta = os.path.join(
            CARPETA_STOCK,
            archivo
        )

        if not os.path.isfile(ruta):
            continue

        nombre = archivo.lower().strip()

        if nombre.endswith(".xlsx") and "stock" in nombre:
            archivos.append(ruta)

    if not archivos:
        raise FileNotFoundError(
            "No se encontró ningún archivo STOCK."
        )

    return max(
        archivos,
        key=os.path.getmtime
    )


# ============================================================
# NORMALIZAR DNI
# ============================================================

def normalizar_dni(valor):

    if valor is None:
        return ""

    texto = str(valor).strip()

    if texto.endswith(".0"):
        texto = texto[:-2]

    texto = "".join(
        caracter
        for caracter in texto
        if caracter.isdigit()
    )

    return texto


# ============================================================
# CONVERTIR FECHA
# ============================================================

def convertir_fecha(valor):

    if valor is None:
        return pd.NaT

    return pd.to_datetime(
        valor,
        dayfirst=True,
        errors="coerce"
    )


# ============================================================
# CARGAR STOCK
#
# IMPORTANTE:
# NO CARGAMOS TODO EL EXCEL EN MEMORIA.
#
# Nos quedamos directamente con:
#
# DNI -> FECHA_ASIG_ESTUDIO MÁS RECIENTE
# ============================================================

def cargar_stock():

    ruta = buscar_ultimo_stock()

    print()
    print("========================================")
    print("STOCK ENCONTRADO")
    print("========================================")
    print(ruta)

    print()
    print("Leyendo STOCK en modo memoria reducida...")

    wb = load_workbook(
        filename=ruta,
        read_only=True,
        data_only=True
    )

    try:

        ws = wb.active

        # ====================================================
        # ENCABEZADOS
        # ====================================================

        encabezados = next(
            ws.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True
            )
        )

        encabezados = [
            str(valor).strip()
            if valor is not None
            else ""
            for valor in encabezados
        ]

        if "NUM_DOC" not in encabezados:

            raise KeyError(
                "No se encontró la columna 'NUM_DOC' en STOCK."
            )

        if "FECHA_ASIG_ESTUDIO" not in encabezados:

            raise KeyError(
                "No se encontró la columna "
                "'FECHA_ASIG_ESTUDIO' en STOCK."
            )

        indice_dni = encabezados.index(
            "NUM_DOC"
        )

        indice_fecha = encabezados.index(
            "FECHA_ASIG_ESTUDIO"
        )

        print()
        print(
            "Columna NUM_DOC:",
            indice_dni + 1
        )

        print(
            "Columna FECHA_ASIG_ESTUDIO:",
            indice_fecha + 1
        )

        # ====================================================
        # DICCIONARIO
        #
        # Solamente guardamos:
        #
        # DNI -> fecha más reciente
        # ====================================================

        stock_dict = {}

        contador = 0

        print()
        print("Procesando registros del STOCK...")

        # ----------------------------------------------------
        # IMPORTANTE:
        # Solo leemos las columnas necesarias.
        # ----------------------------------------------------

        columna_minima = min(
            indice_dni,
            indice_fecha
        ) + 1

        columna_maxima = max(
            indice_dni,
            indice_fecha
        ) + 1

        for fila in ws.iter_rows(
            min_row=2,
            min_col=columna_minima,
            max_col=columna_maxima,
            values_only=True
        ):

            contador += 1

            posicion_dni = (
                indice_dni + 1 - columna_minima
            )

            posicion_fecha = (
                indice_fecha + 1 - columna_minima
            )

            dni = normalizar_dni(
                fila[posicion_dni]
            )

            if not dni:
                continue

            fecha_original = fila[
                posicion_fecha
            ]

            fecha = convertir_fecha(
                fecha_original
            )

            # ------------------------------------------------
            # Si ya existe el DNI, conservar solamente
            # la fecha más reciente.
            # ------------------------------------------------

            if dni not in stock_dict:

                stock_dict[dni] = (
                    fecha_original,
                    fecha
                )

            else:

                fecha_actual = stock_dict[dni][1]

                # Si la nueva fecha es válida y es más reciente
                if (
                    pd.notna(fecha)
                    and (
                        pd.isna(fecha_actual)
                        or fecha > fecha_actual
                    )
                ):

                    stock_dict[dni] = (
                        fecha_original,
                        fecha
                    )

            # ------------------------------------------------
            # Mostrar progreso cada 100.000 registros
            # ------------------------------------------------

            if contador % 100000 == 0:

                print(
                    "Registros procesados:",
                    contador,
                    "| DNI únicos:",
                    len(stock_dict)
                )

        print()
        print(
            "Registros procesados:",
            contador
        )

        print(
            "DNI únicos:",
            len(stock_dict)
        )

    finally:

        wb.close()

    # ========================================================
    # CREAR DATAFRAME PEQUEÑO
    # ========================================================

    documentos = []
    fechas = []

    for dni, valores in stock_dict.items():

        documentos.append(dni)
        fechas.append(valores[0])

    df = pd.DataFrame({
        "NUM_DOC": documentos,
        "FECHA_ASIG_ESTUDIO": fechas
    })

    print()
    print("========================================")
    print("STOCK PROCESADO")
    print("========================================")

    print(
        "Registros finales:",
        len(df)
    )

    return df


# ============================================================
# PREPARAR STOCK
# ============================================================

def preparar_stock(df_stock):

    columnas_necesarias = [
        "NUM_DOC",
        "FECHA_ASIG_ESTUDIO"
    ]

    for columna in columnas_necesarias:

        if columna not in df_stock.columns:

            raise KeyError(
                f"No se encontró la columna "
                f"'{columna}' en STOCK."
            )

    stock = df_stock[
        columnas_necesarias
    ].copy()

    stock["NUM_DOC"] = (
        stock["NUM_DOC"]
        .astype("string")
        .str.strip()
        .str.replace(
            r"\.0$",
            "",
            regex=True
        )
    )

    return stock